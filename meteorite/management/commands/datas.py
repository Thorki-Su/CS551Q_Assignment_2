import os
from pathlib import Path
from django.db import models
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from meteorite.models import Meteorite

class Command(BaseCommand):
    help = 'Load data from csv'

    def handle(self, *args, **options):
        Meteorite.objects.all().delete()
        print('table dropped')

        base_dir = Path(__file__).resolve().parent.parent.parent.parent
        book_path = os.path.join(base_dir, 'meteorite/meteorite_data/meteorites.xlsx')
        try:
            book = load_workbook(book_path)
            sheet = book['meteorites']
        except FileNotFoundError:
            raise CommandError(f"File not found: {book_path}")
        except KeyError:
            raise CommandError("Worksheet named 'meteorites' not found in the Excel file.")
        print(sheet.title)
        max_row_num = sheet.max_row
        max_col_num = sheet.max_column
        print(f'Rows: {max_row_num}, Columns: {max_col_num}')

        
        finish = 0
        
        for i in range(2, max_row_num + 1):
            try:
                row_data = [sheet.cell(row=i, column=j).value for j in range(1, max_col_num + 1)]
                name = row_data[0]
                id = row_data[1]
                nametype = row_data[2]
                recclass = row_data[3]
                mass = float(row_data[4])
                fall = row_data[5]
                year = row_data[6]
                reclat = row_data[7]
                reclong = row_data[8]

                price = 5

                if fall == 'Fell':
                    price += 5
                if nametype == 'Relict':
                    price *= 0.75
                if mass == 0:
                    mass += 0.1
                if mass <= 100:
                    price *= 0.9
                elif mass <= 1000:
                    price = price
                elif mass <= 5000:
                    price *= 1.1
                else:
                    price *= 2

                meteorite = Meteorite.objects.create(
                    name = name,
                    id = id,
                    nametype = nametype,
                    recclass = recclass,
                    mass = mass,
                    fall = fall,
                    year = year,
                    latitude = reclat,
                    longitude = reclong,
                    price = price,
                )
                finish += 1
                print(f'saved, now finish {finish}')
            except Exception as e:
                print(f'Error processing row {i}: {e}')
        print('all data saved')